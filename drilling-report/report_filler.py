"""
Drilling Daily Report Auto-Filler
==================================
Automatically fills a daily drilling report template with data from
GTI report and directional survey files.

Input files (place in the same folder as this script):
    - GTI report:     filename contains "_GTI_report"
    - Survey data:    filename contains "_survey"
    - Report template: filename contains "_daily_report"

Output:
    Template is updated in-place: new row appended with calculated data.
"""

import re
import os
import glob
from datetime import datetime
import pandas as pd
import openpyxl


# ---------------------------------------------------------------------------
# Constants – column numbers in the daily report template
# ---------------------------------------------------------------------------
COL_DATE = 1
COL_MD_FROM = 15
COL_MD_TO = 16
COL_TVD_FROM = 13
COL_TVD_TO = 14
COL_SLIDE_TIME = 24
COL_TOTAL_DRILL_TIME = 25
COL_SLIDE_FOOTAGE = 28
COL_CIRC_TIME = 31
COL_REAMING_TIME = 34
COL_ZENITH_START = 37
COL_ZENITH_END = 38
COL_AZIMUTH_START = 39
COL_AZIMUTH_END = 40
COL_DEVIATION = 41
COL_LOAD = 44
COL_FLOW = 45
COL_RPM = 46
COL_PRESSURE = 47

# Columns in the SLIDE SHEET
SLIDE_COL_DATE = 2
SLIDE_COL_ZENITH = 4
SLIDE_COL_AZIMUTH = 5
SLIDE_COL_TVD = 9
SLIDE_COL_LOAD = 23
SLIDE_COL_RPM = 22
SLIDE_COL_FLOW = 31
SLIDE_COL_PRESSURE = 33
SLIDE_COL_FOOTAGE = 17

# Cell positions in GTI "Summary" sheet
GTI_DATE_ROW, GTI_DATE_COL = 4, 13
GTI_MD_FROM_ROW, GTI_MD_FROM_COL = 6, 14
GTI_MD_TO_ROW, GTI_MD_TO_COL = 6, 20

# Columns that are NOT copied from the previous row
FILLED_COLUMNS = [
    1, 13, 14, 15, 16, 24, 25, 28, 31, 34,
    37, 38, 39, 40, 41, 44, 45, 46, 47,
]


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def find_file(mask):
    """Return the first file matching *mask* in the script directory."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    files = glob.glob(os.path.join(script_dir, mask))
    if not files:
        raise FileNotFoundError(f'File not found: {mask}')
    print(f'Found: {os.path.basename(files[0])}')
    return files[0]


def safe_float(value):
    """Convert *value* to float. Strings like '40-60' are averaged."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        if isinstance(value, str):
            nums = re.findall(r'\d+', value)
            if nums:
                return sum(float(n) for n in nums) / len(nums)
        return None


def write_cell(ws, row, col, value):
    """Write *value* into cell, skipping existing formulas."""
    if col is None or value is None:
        return
    cell = ws.cell(row=row, column=col)
    if cell.value and isinstance(cell.value, str) and str(cell.value).startswith('='):
        return
    cell.value = value


def parse_date(cell_value):
    """Parse a date from an Excel cell value."""
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, str):
        return pd.to_datetime(cell_value, dayfirst=True).date()
    raise ValueError(f'Cannot parse date: {cell_value}')


def copy_formulas_and_values(ws, source_row, target_row):
    """Copy all cells from *source_row* to *target_row*.

    Formula row references are updated automatically.
    Columns listed in FILLED_COLUMNS are skipped.
    """
    for col in range(1, ws.max_column + 1):
        if col in FILLED_COLUMNS:
            continue
        source = ws.cell(row=source_row, column=col).value
        if source is None:
            continue
        if isinstance(source, str) and source.startswith('='):
            def replace_row(m):
                letters, digits = m.group(1), m.group(2)
                return letters + str(target_row) if int(digits) == source_row else m.group(0)
            new_formula = re.sub(r'([A-Z]+)(\d+)', replace_row, source)
            ws.cell(row=target_row, column=col).value = new_formula
        else:
            ws.cell(row=target_row, column=col).value = source


# ---------------------------------------------------------------------------
# 1. GTI report
# ---------------------------------------------------------------------------

def read_gti(path):
    """Extract date, MD interval, drilling times from GTI report."""
    print('\n' + '=' * 60)
    print('1. GTI REPORT')
    print('=' * 60)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Сводка ГТИ']

    report_date = parse_date(ws.cell(row=GTI_DATE_ROW, column=GTI_DATE_COL).value)
    md_from = float(ws.cell(row=GTI_MD_FROM_ROW, column=GTI_MD_FROM_COL).value)
    md_to = float(ws.cell(row=GTI_MD_TO_ROW, column=GTI_MD_TO_COL).value)
    print(f'   Date: {report_date}  |  MD: {md_from} – {md_to} m')

    rotary = slide = circ = reaming = survey = 0.0
    for row in range(1, ws.max_row + 1):
        for col in range(1, 30):
            if 'углубление скважины' in str(ws.cell(row=row, column=col).value or '').lower():
                for r in range(row + 1, row + 5):
                    vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
                    if sum(1 for v in vals if isinstance(v, (int, float))) >= 5:
                        rotary = float(vals[0]) if isinstance(vals[0], (int, float)) else 0
                        slide = float(vals[1]) if isinstance(vals[1], (int, float)) else 0
                        circ = float(vals[3]) if isinstance(vals[3], (int, float)) else 0
                        reaming = float(vals[4]) if isinstance(vals[4], (int, float)) else 0
                        survey = float(vals[6]) if isinstance(vals[6], (int, float)) else 0
                        break
                break
        if rotary > 0:
            break

    total_drill = rotary + slide
    circ_total = total_drill + circ + reaming + survey
    print(f'   Rotary: {rotary:.2f}  Slide: {slide:.2f}  Circ: {circ_total:.2f}  Reaming: {reaming:.2f}')
    return {
        'date': report_date, 'md_from': md_from, 'md_to': md_to,
        'rotary_time': rotary, 'slide_time': slide,
        'circ_time': circ_total, 'reaming_time': reaming,
        'total_drill_time': total_drill,
    }


# ---------------------------------------------------------------------------
# 2. SLIDE SHEET
# ---------------------------------------------------------------------------

def read_slidesheet(path, md_from, target_date):
    """Extract directional data and average drilling parameters."""
    print('\n' + '=' * 60)
    print('2. SLIDE SHEET')
    print('=' * 60)
    wb = openpyxl.load_workbook(path, data_only=True)

    ws = None
    for name in wb.sheetnames:
        if 'slide sheet' in name.lower():
            temp = wb[name]
            try:
                a = float(temp.cell(row=5, column=4).value or 0)
                b = float(temp.cell(row=5, column=7).value or 0)
                if a <= md_from <= b:
                    ws = temp
                    break
            except (ValueError, TypeError):
                pass
    if ws is None:
        for name in wb.sheetnames:
            if 'slide sheet' in name.lower():
                ws = wb[name]
                break
    if ws is None:
        raise ValueError('SLIDE SHEET not found')

    print(f'   Sheet: {ws.title}')
    rows, cur_date = [], None
    for row in range(13, ws.max_row + 1):
        dv = ws.cell(row=row, column=SLIDE_COL_DATE).value
        if dv:
            try:
                cur_date = parse_date(dv)
            except ValueError:
                pass
        if cur_date is None:
            continue
        rows.append({
            'date': cur_date,
            'zenith': safe_float(ws.cell(row=row, column=SLIDE_COL_ZENITH).value),
            'azimuth': safe_float(ws.cell(row=row, column=SLIDE_COL_AZIMUTH).value),
            'tvd': safe_float(ws.cell(row=row, column=SLIDE_COL_TVD).value),
            'load': safe_float(ws.cell(row=row, column=SLIDE_COL_LOAD).value),
            'rpm': safe_float(ws.cell(row=row, column=SLIDE_COL_RPM).value),
            'flow': safe_float(ws.cell(row=row, column=SLIDE_COL_FLOW).value),
            'pressure': safe_float(ws.cell(row=row, column=SLIDE_COL_PRESSURE).value),
            'footage': safe_float(ws.cell(row=row, column=SLIDE_COL_FOOTAGE).value),
        })

    df = pd.DataFrame(rows)
    df_day = df[df['date'] == target_date]
    print(f'   Rows for {target_date}: {len(df_day)}')
    if df_day.empty:
        return {}

    df_angle = df_day.dropna(subset=['zenith'])
    if len(df_angle) >= 2:
        zs, ze = df_angle.iloc[0]['zenith'], df_angle.iloc[-1]['zenith']
        azs, aze = df_angle.iloc[0]['azimuth'], df_angle.iloc[-1]['azimuth']
        ts, te = df_angle.iloc[0]['tvd'], df_angle.iloc[-1]['tvd']
    elif len(df_angle) == 1:
        zs = ze = df_angle.iloc[0]['zenith']
        azs = aze = df_angle.iloc[0]['azimuth']
        ts = te = df_angle.iloc[0]['tvd']
    else:
        zs = ze = azs = aze = ts = te = None

    avg_load = df_day['load'].dropna().mean()
    avg_flow = df_day['flow'].dropna().mean()
    avg_rpm = df_day['rpm'].dropna().mean()
    avg_pressure = df_day['pressure'].dropna().mean()
    footage = df_day['footage'].dropna().sum()

    print(f'   Zen: {zs}→{ze}  Az: {azs}→{aze}  TVD: {ts}→{te}')
    print(f'   G={avg_load:.1f}  Q={avg_flow:.1f}  N={avg_rpm:.0f}  P={avg_pressure:.1f}  Slide: {footage:.2f} m')
    return {
        'zenith_start': zs, 'zenith_end': ze,
        'azimuth_start': azs, 'azimuth_end': aze,
        'tvd_start': ts, 'tvd_end': te,
        'avg_load': avg_load, 'avg_flow': avg_flow,
        'avg_rpm': avg_rpm, 'avg_pressure': avg_pressure,
        'slide_footage': footage,
    }


# ---------------------------------------------------------------------------
# 3. Deviation from vertical
# ---------------------------------------------------------------------------

def read_deviation(path, md_to):
    """Return deviation-from-vertical closest to *md_to* (sheet 'Замеры')."""
    print('\n' + '=' * 60)
    print('3. DEVIATION')
    print('=' * 60)
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Замеры' not in wb.sheetnames:
        print('   Sheet "Замеры" not found')
        return 0.0
    ws = wb['Замеры']

    # Find header row
    hdr = None
    for row in range(1, 50):
        txt = ' '.join([str(ws.cell(row=row, column=c).value or '') for c in range(1, 20)])
        if 'глубин' in txt.lower() and 'зенит' in txt.lower():
            hdr = row
            break
    if hdr is None:
        print('   Header not found')
        return 0.0

    headers = {}
    for col in range(1, 20):
        v = ws.cell(row=hdr, column=col).value
        if v:
            headers[str(v).lower().strip()] = col

    depth_col = next((c for k, c in headers.items() if 'глуб' in k and ('ствол' in k or 'md' in k)), 1)
    dev_col = 11
    if not (ws.cell(row=hdr, column=11).value and 'отклон' in str(ws.cell(row=hdr, column=11).value).lower()):
        dev_col = next((c for k, c in headers.items() if 'отклон' in k or 'отход' in k), 11)

    depths = []
    for row in range(hdr + 1, ws.max_row + 1):
        v = ws.cell(row=row, column=depth_col).value
        if v:
            try:
                depths.append((float(v), row))
            except (ValueError, TypeError):
                pass
    if not depths:
        print('   No depths found')
        return 0.0

    closest = min(depths, key=lambda x: abs(x[0] - md_to))
    dev = float(ws.cell(row=closest[1], column=dev_col).value or 0)
    print(f'   MD end: {md_to}  Closest: {closest[0]}  Deviation: {dev} m')
    return dev


# ---------------------------------------------------------------------------
# 4. Fill the daily report template
# ---------------------------------------------------------------------------

def fill_report(path, gti, slide, deviation):
    """Populate the daily report template with all collected data."""
    print('\n' + '=' * 60)
    print('4. FILLING REPORT')
    print('=' * 60)
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    target = None
    for row in range(14, 200):
        if ws.cell(row=row, column=1).value is None:
            target = row
            break
    target = target or 14
    print(f'   Row: {target}')

    write_cell(ws, target, COL_DATE, gti['date'])
    write_cell(ws, target, COL_MD_FROM, gti['md_from'])
    write_cell(ws, target, COL_MD_TO, gti['md_to'])
    write_cell(ws, target, COL_SLIDE_TIME, round(gti['slide_time'], 2))
    write_cell(ws, target, COL_TOTAL_DRILL_TIME, round(gti['total_drill_time'], 2))
    write_cell(ws, target, COL_CIRC_TIME, round(gti['circ_time'], 2))
    write_cell(ws, target, COL_REAMING_TIME, round(gti['reaming_time'], 2))

    if slide.get('tvd_start') is not None:
        write_cell(ws, target, COL_TVD_FROM, round(slide['tvd_start'], 2))
        write_cell(ws, target, COL_TVD_TO, round(slide['tvd_end'], 2))
    if slide.get('zenith_start') is not None:
        write_cell(ws, target, COL_ZENITH_START, round(slide['zenith_start'], 2))
        write_cell(ws, target, COL_ZENITH_END, round(slide['zenith_end'], 2))
        write_cell(ws, target, COL_AZIMUTH_START, round(slide['azimuth_start'], 2))
        write_cell(ws, target, COL_AZIMUTH_END, round(slide['azimuth_end'], 2))

    write_cell(ws, target, COL_SLIDE_FOOTAGE, round(slide.get('slide_footage', 0), 2))
    write_cell(ws, target, COL_DEVIATION, round(deviation, 2) if deviation else None)

    for key, col in [('avg_load', COL_LOAD), ('avg_flow', COL_FLOW),
                     ('avg_rpm', COL_RPM), ('avg_pressure', COL_PRESSURE)]:
        val = slide.get(key)
        if val is not None and not pd.isna(val):
            decimals = 0 if key == 'avg_rpm' else 1
            write_cell(ws, target, col, round(val, decimals))

    if target > 14:
        copy_formulas_and_values(ws, target - 1, target)

    wb.save(path)
    print('\n' + '=' * 60)
    print('DONE!')
    print('=' * 60)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    gti = read_gti(find_file('*_GTI_report*.xlsx'))
    survey_path = find_file('*_survey*.xlsx')
    slide = read_slidesheet(survey_path, gti['md_from'], gti['date'])
    dev = read_deviation(survey_path, gti['md_to'])
    fill_report(find_file('*_daily_report*.xlsx'), gti, slide, dev)


if __name__ == '__main__':
    main()